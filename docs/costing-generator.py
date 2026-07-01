#!/usr/bin/env python3
"""Generate costing Excel for Agent Onboarding System - Amended Pricing"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ─── Styles ───────────────────────────────────────────────
header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
section_font = Font(name='Calibri', bold=True, size=11, color='1F4E79')
section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
total_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
total_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
grand_total_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
margin_good_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
margin_font = Font(name='Calibri', bold=True, size=11, color='375623')
normal_font = Font(name='Calibri', size=11)
bold_font = Font(name='Calibri', bold=True, size=11)
quota_font = Font(name='Calibri', size=10, italic=True, color='1F4E79')
currency_format = '#,##0'
pct_format = '0%'
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

def style_section(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border

def style_total(ws, row, cols, fill=total_fill):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = total_font
        cell.fill = fill
        cell.border = thin_border

def style_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.border = thin_border

def style_margin_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = margin_font
        cell.fill = margin_good_fill
        cell.border = thin_border

def right_align(ws, row, cols_list):
    for c in cols_list:
        ws.cell(row=row, column=c).alignment = Alignment(horizontal='right')


# ═══════════════════════════════════════════════════════════
# SHEET 1: Option A - Managed SaaS (Amended)
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Option A - Managed SaaS"
ws1.sheet_properties.tabColor = "1F4E79"

ws1.column_dimensions['A'].width = 5
ws1.column_dimensions['B'].width = 55
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 18

# Title
ws1.merge_cells('A1:D1')
ws1['A1'].value = "AGENT ONBOARDING SYSTEM - MONTHLY SaaS SUBSCRIPTION"
ws1['A1'].font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:D2')
ws1['A2'].value = "All prices in MYR (Malaysian Ringgit) | Monthly Lump Sum with Usage Quotas"
ws1['A2'].font = Font(name='Calibri', italic=True, size=10, color='808080')
ws1['A2'].alignment = Alignment(horizontal='center')

# Headers
row = 4
ws1.cell(row=row, column=1, value="#")
ws1.cell(row=row, column=2, value="Description")
ws1.cell(row=row, column=3, value="Monthly (RM)")
ws1.cell(row=row, column=4, value="Annual (RM)")
style_header(ws1, row, 4)
ws1.row_dimensions[row].height = 25

# ─── A: SaaS Platform ───
row = 5
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="A. SaaS PLATFORM")
style_section(ws1, row, 4)

platform_items = [
    (1, "Admin Portal - Event management, agent/unit management, tier & rewards config, reports", 300),
    (2, "Agent Portal - Campaign browsing, link generation & sharing, partner management, rewards", 300),
    (3, "Public Pages - Registration, QR check-in display, OTP checkout, feedback", 300),
]

platform_start = row + 1
for i, (num, desc, price) in enumerate(platform_items):
    row = platform_start + i
    ws1.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws1.cell(row=row, column=2, value=desc)
    ws1.cell(row=row, column=3, value=price).number_format = currency_format
    ws1.cell(row=row, column=4).value = f"=C{row}*12"
    ws1.cell(row=row, column=4).number_format = currency_format
    style_row(ws1, row, 4)
    right_align(ws1, row, [3, 4])

row = platform_start + len(platform_items)
ws1.cell(row=row, column=2, value="Subtotal - SaaS Platform")
ws1.cell(row=row, column=3).value = f"=SUM(C{platform_start}:C{row-1})"
ws1.cell(row=row, column=3).number_format = currency_format
ws1.cell(row=row, column=4).value = f"=C{row}*12"
ws1.cell(row=row, column=4).number_format = currency_format
style_total(ws1, row, 4)
right_align(ws1, row, [3, 4])
sub_a = row

# ─── B: Cloud Infra & Services ───
row += 1
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="B. CLOUD INFRASTRUCTURE & SERVICES (BUNDLED WITH QUOTA)")
style_section(ws1, row, 4)

# Item 4-6 are monthly, item 7 is per year
infra_monthly_items = [
    (4, "Cloud Hosting - 3 production apps, CI/CD auto-deployment, SSL", 200, "Quota: 3 apps, 100GB bandwidth/mo"),
    (5, "Database & Backend - PostgreSQL, Auth, Edge Functions, automated backups", 250, "Quota: 8GB database, 500MB storage"),
    (6, "Email Service - Event reminders, notifications, transactional emails (Resend)", 150, "Quota: 5,000 emails/mo"),
]

infra_start = row + 1
infra_price_rows = []
for i, (num, desc, price, quota) in enumerate(infra_monthly_items):
    row = infra_start + (i * 2)  # each item takes 2 rows (item + quota)
    ws1.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws1.cell(row=row, column=2, value=desc)
    ws1.cell(row=row, column=3, value=price).number_format = currency_format
    ws1.cell(row=row, column=4).value = f"=C{row}*12"
    ws1.cell(row=row, column=4).number_format = currency_format
    style_row(ws1, row, 4)
    right_align(ws1, row, [3, 4])
    infra_price_rows.append(row)
    # quota sub-row
    row += 1
    ws1.cell(row=row, column=2, value=f"     {quota}")
    ws1.cell(row=row, column=2).font = quota_font
    style_row(ws1, row, 4)

# Domain - per year item
row += 1
ws1.cell(row=row, column=1, value=7).alignment = Alignment(horizontal='center')
ws1.cell(row=row, column=2, value="Domain & SSL Management (Per Year)")
ws1.cell(row=row, column=3, value=250).number_format = currency_format
ws1.cell(row=row, column=4, value="")  # no annual calc, it IS the annual
style_row(ws1, row, 4)
right_align(ws1, row, [3])
domain_row = row
infra_price_rows.append(row)

# Subtotal B
row += 1
ws1.cell(row=row, column=2, value="Subtotal - Cloud Infrastructure & Services")
# Sum only the price rows
sum_formula = "+".join([f"C{r}" for r in infra_price_rows])
ws1.cell(row=row, column=3).value = f"={sum_formula}"
ws1.cell(row=row, column=3).number_format = currency_format
# Annual: monthly items * 12 + domain (already annual)
monthly_sum = "+".join([f"C{r}" for r in infra_price_rows[:-1]])  # exclude domain
ws1.cell(row=row, column=4).value = f"=({monthly_sum})*12+C{domain_row}"
ws1.cell(row=row, column=4).number_format = currency_format
style_total(ws1, row, 4)
right_align(ws1, row, [3, 4])
sub_b = row

# ─── C: Support & SLA ───
row += 1
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="C. SUPPORT & SLA")
style_section(ws1, row, 4)

support_start = row + 1
support_items = [
    (8, "SLA Support - 99.5% uptime guarantee, P1-P4 incident response", 300),
    (9, "Bug Fixes, System Monitoring & Maintenance", 200),
    (10, "Minor Adjustments & Updates (up to 2 hours/month)", 200),
]

for i, (num, desc, price) in enumerate(support_items):
    row = support_start + i
    ws1.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws1.cell(row=row, column=2, value=desc)
    ws1.cell(row=row, column=3, value=price).number_format = currency_format
    ws1.cell(row=row, column=4).value = f"=C{row}*12"
    ws1.cell(row=row, column=4).number_format = currency_format
    style_row(ws1, row, 4)
    right_align(ws1, row, [3, 4])

row = support_start + len(support_items)
ws1.cell(row=row, column=2, value="Subtotal - Support & SLA")
ws1.cell(row=row, column=3).value = f"=SUM(C{support_start}:C{row-1})"
ws1.cell(row=row, column=3).number_format = currency_format
ws1.cell(row=row, column=4).value = f"=C{row}*12"
ws1.cell(row=row, column=4).number_format = currency_format
style_total(ws1, row, 4)
right_align(ws1, row, [3, 4])
sub_c = row

# ─── D: Training ───
row += 1
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="D. TRAINING & ENABLEMENT")
style_section(ws1, row, 4)

training_start = row + 1
training_items = [
    (11, "New Agent/Unit Onboarding Sessions (1 session/month, up to 20 pax)", 200),
]

for i, (num, desc, price) in enumerate(training_items):
    row = training_start + i
    ws1.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws1.cell(row=row, column=2, value=desc)
    ws1.cell(row=row, column=3, value=price).number_format = currency_format
    ws1.cell(row=row, column=4).value = f"=C{row}*12"
    ws1.cell(row=row, column=4).number_format = currency_format
    style_row(ws1, row, 4)
    right_align(ws1, row, [3, 4])

row = training_start + len(training_items)
ws1.cell(row=row, column=2, value="Subtotal - Training & Enablement")
ws1.cell(row=row, column=3).value = f"=SUM(C{training_start}:C{row-1})"
ws1.cell(row=row, column=3).number_format = currency_format
ws1.cell(row=row, column=4).value = f"=C{row}*12"
ws1.cell(row=row, column=4).number_format = currency_format
style_total(ws1, row, 4)
right_align(ws1, row, [3, 4])
sub_d = row

# ─── GRAND TOTAL ───
row += 2
ws1.cell(row=row, column=2, value="TOTAL MONTHLY LUMP SUM")
ws1.cell(row=row, column=3).value = f"=C{sub_a}+C{sub_b}+C{sub_c}+C{sub_d}"
ws1.cell(row=row, column=3).number_format = currency_format
ws1.cell(row=row, column=4).value = f"=C{row}*12"
ws1.cell(row=row, column=4).number_format = currency_format
style_total(ws1, row, 4, grand_total_fill)
right_align(ws1, row, [3, 4])
ws1.row_dimensions[row].height = 30
grand_total_row_s1 = row

# ─── Quotas ───
row += 2
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="INCLUDED MONTHLY QUOTAS")
style_section(ws1, row, 4)

quotas = [
    ("Cloud Hosting", "3 production apps, 100GB bandwidth/mo, CI/CD auto-deploy"),
    ("Database & Storage", "8GB PostgreSQL database, 500MB file storage, automated daily backups"),
    ("Email (Resend)", "5,000 transactional emails/month"),
    ("SSL & Domain", "Managed SSL certificates, custom domain support"),
    ("Support Hours", "2 hours/month minor adjustments included"),
    ("Training", "1 group onboarding session/month (up to 20 pax)"),
]

for service, quota in quotas:
    row += 1
    ws1.cell(row=row, column=2, value=service).font = bold_font
    ws1.merge_cells(f'C{row}:D{row}')
    ws1.cell(row=row, column=3, value=quota).font = normal_font
    style_row(ws1, row, 4)

# ─── Overage Rates ───
row += 2
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="OVERAGE RATES (BEYOND MONTHLY QUOTA)")
style_section(ws1, row, 4)

overages = [
    ("Additional emails (beyond 5,000/mo)", "RM 0.01/email"),
    ("Additional OTP verifications (beyond 500/mo)", "RM 0.30/OTP"),
    ("Database storage upgrade (beyond 8GB)", "RM 50/GB/mo"),
    ("Additional support hours (beyond 2 hrs/mo)", "RM 150/hour"),
    ("Additional training sessions", "RM 500/session"),
]

for desc, rate in overages:
    row += 1
    ws1.cell(row=row, column=2, value=desc)
    ws1.merge_cells(f'C{row}:D{row}')
    ws1.cell(row=row, column=3, value=rate).font = bold_font
    ws1.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    style_row(ws1, row, 4)

# ─── Optional Add-Ons ───
row += 2
ws1.merge_cells(f'A{row}:D{row}')
ws1.cell(row=row, column=1, value="OPTIONAL ADD-ONS (QUOTED SEPARATELY)")
style_section(ws1, row, 4)

addon_items = [
    ("New Feature Development", "RM 150/hour"),
    ("Custom Integrations (CRM, HR, ERP)", "Quoted per scope"),
]

for desc, rate in addon_items:
    row += 1
    ws1.cell(row=row, column=2, value=desc)
    ws1.merge_cells(f'C{row}:D{row}')
    ws1.cell(row=row, column=3, value=rate)
    ws1.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    style_row(ws1, row, 4)


# ═══════════════════════════════════════════════════════════
# SHEET 2: Internal Cost & Margin
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Internal Cost & Margin")
ws2.sheet_properties.tabColor = "375623"

ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 50
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 14

# Title
ws2.merge_cells('A1:F1')
ws2['A1'].value = "INTERNAL COST ANALYSIS & PROFIT MARGIN"
ws2['A1'].font = Font(name='Calibri', bold=True, size=14, color='375623')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

ws2.merge_cells('A2:F2')
ws2['A2'].value = "CONFIDENTIAL - For Your Eyes Only | All prices in MYR"
ws2['A2'].font = Font(name='Calibri', italic=True, size=10, color='C00000')
ws2['A2'].alignment = Alignment(horizontal='center')

# Headers
row = 4
for i, h in enumerate(["#", "Description", "Client Pays (RM)", "Your Cost (RM)", "Margin (RM)", "Margin %"], 1):
    ws2.cell(row=row, column=i, value=h)
style_header(ws2, row, 6)
ws2.row_dimensions[row].height = 30

# ─── A: Platform (pure margin) ───
row = 5
ws2.merge_cells(f'A{row}:F{row}')
ws2.cell(row=row, column=1, value="A. SaaS PLATFORM (SOFTWARE - ALREADY BUILT)")
style_section(ws2, row, 6)

p_start = row + 1
platform_cost = [
    (1, "Admin Portal", 300, 0),
    (2, "Agent Portal", 300, 0),
    (3, "Public Pages", 300, 0),
]

for i, (num, desc, client, cost) in enumerate(platform_cost):
    row = p_start + i
    ws2.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=client).number_format = currency_format
    ws2.cell(row=row, column=4, value=cost).number_format = currency_format
    ws2.cell(row=row, column=5).value = f"=C{row}-D{row}"
    ws2.cell(row=row, column=5).number_format = currency_format
    ws2.cell(row=row, column=6).value = f"=IF(C{row}=0,0,E{row}/C{row})"
    ws2.cell(row=row, column=6).number_format = pct_format
    style_row(ws2, row, 6)
    right_align(ws2, row, [3, 4, 5, 6])

row = p_start + len(platform_cost)
ws2.cell(row=row, column=2, value="Subtotal - Platform")
ws2.cell(row=row, column=3).value = f"=SUM(C{p_start}:C{row-1})"
ws2.cell(row=row, column=3).number_format = currency_format
ws2.cell(row=row, column=4).value = f"=SUM(D{p_start}:D{row-1})"
ws2.cell(row=row, column=4).number_format = currency_format
ws2.cell(row=row, column=5).value = f"=C{row}-D{row}"
ws2.cell(row=row, column=5).number_format = currency_format
ws2.cell(row=row, column=6).value = f"=IF(C{row}=0,0,E{row}/C{row})"
ws2.cell(row=row, column=6).number_format = pct_format
style_total(ws2, row, 6)
right_align(ws2, row, [3, 4, 5, 6])
sub_a2 = row

# ─── B: Infra (client pays bundled, you have real cost) ───
row += 1
ws2.merge_cells(f'A{row}:F{row}')
ws2.cell(row=row, column=1, value="B. CLOUD INFRASTRUCTURE & SERVICES (BUNDLED - YOUR REAL COST)")
style_section(ws2, row, 6)

infra_start2 = row + 1
infra_cost = [
    (4, "Render - Cloud Hosting (3 Static Sites + CI/CD)", 200, 50),
    (5, "Supabase - Database, Auth, Edge Functions, Storage", 250, 110),
    (6, "Resend - Email Service (5,000 emails/mo quota)", 150, 90),
    (7, "Domain & SSL (Per Year - amortized monthly)", 21, 15),
]
# Note: item 7 client pays RM 250/yr = ~RM 21/mo amortized for comparison

for i, (num, desc, client, cost) in enumerate(infra_cost):
    row = infra_start2 + i
    ws2.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=client).number_format = currency_format
    ws2.cell(row=row, column=4, value=cost).number_format = currency_format
    ws2.cell(row=row, column=5).value = f"=C{row}-D{row}"
    ws2.cell(row=row, column=5).number_format = currency_format
    ws2.cell(row=row, column=6).value = f"=IF(C{row}=0,0,E{row}/C{row})"
    ws2.cell(row=row, column=6).number_format = pct_format
    style_row(ws2, row, 6)
    right_align(ws2, row, [3, 4, 5, 6])

row = infra_start2 + len(infra_cost)
ws2.cell(row=row, column=2, value="Subtotal - Infra & Services (monthly equivalent)")
ws2.cell(row=row, column=3).value = f"=SUM(C{infra_start2}:C{row-1})"
ws2.cell(row=row, column=3).number_format = currency_format
ws2.cell(row=row, column=4).value = f"=SUM(D{infra_start2}:D{row-1})"
ws2.cell(row=row, column=4).number_format = currency_format
ws2.cell(row=row, column=5).value = f"=C{row}-D{row}"
ws2.cell(row=row, column=5).number_format = currency_format
ws2.cell(row=row, column=6).value = f"=IF(C{row}=0,0,E{row}/C{row})"
ws2.cell(row=row, column=6).number_format = pct_format
style_total(ws2, row, 6)
right_align(ws2, row, [3, 4, 5, 6])
sub_b2 = row

# ─── C: Support (your time) ───
row += 1
ws2.merge_cells(f'A{row}:F{row}')
ws2.cell(row=row, column=1, value="C. SUPPORT & SLA (YOUR TIME)")
style_section(ws2, row, 6)

sup_start = row + 1
support_cost = [
    (8, "SLA Support", 300, 0),
    (9, "Bug Fixes & Monitoring", 200, 0),
    (10, "Minor Adjustments (2 hrs/mo)", 200, 0),
]

for i, (num, desc, client, cost) in enumerate(support_cost):
    row = sup_start + i
    ws2.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=client).number_format = currency_format
    ws2.cell(row=row, column=4, value=cost).number_format = currency_format
    ws2.cell(row=row, column=5).value = f"=C{row}-D{row}"
    ws2.cell(row=row, column=5).number_format = currency_format
    ws2.cell(row=row, column=6).value = f"=IF(C{row}=0,0,E{row}/C{row})"
    ws2.cell(row=row, column=6).number_format = pct_format
    style_row(ws2, row, 6)
    right_align(ws2, row, [3, 4, 5, 6])

row = sup_start + len(support_cost)
ws2.cell(row=row, column=2, value="Subtotal - Support")
ws2.cell(row=row, column=3).value = f"=SUM(C{sup_start}:C{row-1})"
ws2.cell(row=row, column=3).number_format = currency_format
ws2.cell(row=row, column=4).value = f"=SUM(D{sup_start}:D{row-1})"
ws2.cell(row=row, column=4).number_format = currency_format
ws2.cell(row=row, column=5).value = f"=C{row}-D{row}"
ws2.cell(row=row, column=5).number_format = currency_format
ws2.cell(row=row, column=6).value = f"=IF(C{row}=0,0,E{row}/C{row})"
ws2.cell(row=row, column=6).number_format = pct_format
style_total(ws2, row, 6)
right_align(ws2, row, [3, 4, 5, 6])
sub_c2 = row

# ─── D: Training (your time) ───
row += 1
ws2.merge_cells(f'A{row}:F{row}')
ws2.cell(row=row, column=1, value="D. TRAINING (YOUR TIME)")
style_section(ws2, row, 6)

train_start = row + 1
training_cost = [
    (11, "Agent Onboarding (1 session/mo)", 200, 0),
]

for i, (num, desc, client, cost) in enumerate(training_cost):
    row = train_start + i
    ws2.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws2.cell(row=row, column=2, value=desc)
    ws2.cell(row=row, column=3, value=client).number_format = currency_format
    ws2.cell(row=row, column=4, value=cost).number_format = currency_format
    ws2.cell(row=row, column=5).value = f"=C{row}-D{row}"
    ws2.cell(row=row, column=5).number_format = currency_format
    ws2.cell(row=row, column=6).value = f"=IF(C{row}=0,0,E{row}/C{row})"
    ws2.cell(row=row, column=6).number_format = pct_format
    style_row(ws2, row, 6)
    right_align(ws2, row, [3, 4, 5, 6])

row = train_start + len(training_cost)
ws2.cell(row=row, column=2, value="Subtotal - Training")
ws2.cell(row=row, column=3).value = f"=SUM(C{train_start}:C{row-1})"
ws2.cell(row=row, column=3).number_format = currency_format
ws2.cell(row=row, column=4).value = f"=SUM(D{train_start}:D{row-1})"
ws2.cell(row=row, column=4).number_format = currency_format
ws2.cell(row=row, column=5).value = f"=C{row}-D{row}"
ws2.cell(row=row, column=5).number_format = currency_format
ws2.cell(row=row, column=6).value = f"=IF(C{row}=0,0,E{row}/C{row})"
ws2.cell(row=row, column=6).number_format = pct_format
style_total(ws2, row, 6)
right_align(ws2, row, [3, 4, 5, 6])
sub_d2 = row

# ─── MONTHLY TOTALS ───
row += 2
ws2.cell(row=row, column=2, value="TOTAL MONTHLY REVENUE (CLIENT PAYS)")
ws2.cell(row=row, column=3).value = f"=C{sub_a2}+C{sub_b2}+C{sub_c2}+C{sub_d2}"
ws2.cell(row=row, column=3).number_format = currency_format
style_total(ws2, row, 6, grand_total_fill)
right_align(ws2, row, [3, 4, 5, 6])
ws2.row_dimensions[row].height = 28
rev_row = row

row += 1
ws2.cell(row=row, column=2, value="TOTAL MONTHLY HARD COST (Supabase + Render + Resend + Domain)")
ws2.cell(row=row, column=4).value = f"=D{sub_b2}"
ws2.cell(row=row, column=4).number_format = currency_format
ws2.cell(row=row, column=4).font = Font(name='Calibri', bold=True, color='C00000')
style_total(ws2, row, 6)
right_align(ws2, row, [3, 4, 5, 6])
cost_row = row

row += 1
ws2.cell(row=row, column=2, value="MONTHLY GROSS PROFIT (before your time)")
ws2.cell(row=row, column=5).value = f"=C{rev_row}-D{cost_row}"
ws2.cell(row=row, column=5).number_format = currency_format
ws2.cell(row=row, column=6).value = f"=IF(C{rev_row}=0,0,E{row}/C{rev_row})"
ws2.cell(row=row, column=6).number_format = pct_format
style_margin_row(ws2, row, 6)
ws2.cell(row=row, column=2).font = Font(name='Calibri', bold=True, size=12, color='375623')
ws2.cell(row=row, column=5).font = Font(name='Calibri', bold=True, size=12, color='375623')
right_align(ws2, row, [5, 6])
ws2.row_dimensions[row].height = 28
gross_row = row

# ─── Your Time Cost ───
row += 2
ws2.merge_cells(f'A{row}:F{row}')
ws2.cell(row=row, column=1, value="YOUR TIME INVESTMENT (EDITABLE)")
style_section(ws2, row, 6)

row += 1
ws2.cell(row=row, column=2, value="Estimated hours/month (support + training + monitoring)")
ws2.cell(row=row, column=4, value=6)
style_row(ws2, row, 6)
ws2.cell(row=row, column=2).font = Font(name='Calibri', italic=True, size=10)
right_align(ws2, row, [4])
hrs_row = row

row += 1
ws2.cell(row=row, column=2, value="Your hourly rate (RM)")
ws2.cell(row=row, column=4, value=100).number_format = currency_format
style_row(ws2, row, 6)
ws2.cell(row=row, column=2).font = Font(name='Calibri', italic=True, size=10)
right_align(ws2, row, [4])
rate_row = row

row += 1
ws2.cell(row=row, column=2, value="Monthly time cost").font = bold_font
ws2.cell(row=row, column=4).value = f"=D{hrs_row}*D{rate_row}"
ws2.cell(row=row, column=4).number_format = currency_format
style_row(ws2, row, 6)
right_align(ws2, row, [4])
time_cost = row

row += 1
ws2.cell(row=row, column=2, value="MONTHLY NET PROFIT (after infra + your time)")
ws2.cell(row=row, column=5).value = f"=E{gross_row}-D{time_cost}"
ws2.cell(row=row, column=5).number_format = currency_format
ws2.cell(row=row, column=6).value = f"=IF(C{rev_row}=0,0,E{row}/C{rev_row})"
ws2.cell(row=row, column=6).number_format = pct_format
style_margin_row(ws2, row, 6)
ws2.cell(row=row, column=2).font = Font(name='Calibri', bold=True, size=11, color='375623')
right_align(ws2, row, [5, 6])
ws2.row_dimensions[row].height = 28
net_row = row

row += 1
ws2.cell(row=row, column=2, value="ANNUAL NET PROFIT (12-month contract)")
ws2.cell(row=row, column=5).value = f"=E{net_row}*12"
ws2.cell(row=row, column=5).number_format = currency_format
style_margin_row(ws2, row, 6)
ws2.cell(row=row, column=2).font = Font(name='Calibri', bold=True, size=12, color='375623')
ws2.cell(row=row, column=5).font = Font(name='Calibri', bold=True, size=12, color='375623')
right_align(ws2, row, [5])

# ─── Infra Reference ───
row += 2
ws2.merge_cells(f'A{row}:F{row}')
ws2.cell(row=row, column=1, value="INFRASTRUCTURE COST REFERENCE (USD \u2192 MYR @ ~4.40)")
style_section(ws2, row, 6)

ref_items = [
    ("Supabase Pro Plan", "USD 25/mo", "~RM 110/mo", "Database, Auth, Edge Functions, Storage, Realtime"),
    ("Render (3 Static Sites)", "Free - USD 12/mo", "~RM 0-50/mo", "Free tier likely sufficient; paid for custom domains"),
    ("Resend (Email Pro)", "USD 20/mo", "~RM 90/mo", "Up to 50,000 emails/mo; client quota 5,000"),
    ("Domain (.com)", "~USD 12/yr", "~RM 15/mo amort.", "Annual cost split monthly"),
]

for desc, usd, myr, notes in ref_items:
    row += 1
    ws2.cell(row=row, column=2, value=desc).font = bold_font
    ws2.cell(row=row, column=3, value=usd)
    ws2.cell(row=row, column=4, value=myr)
    ws2.merge_cells(f'E{row}:F{row}')
    ws2.cell(row=row, column=5, value=notes).font = Font(name='Calibri', size=9, color='808080')
    style_row(ws2, row, 6)

# ─── Notes ───
row += 2
ws2.merge_cells(f'A{row}:F{row}')
ws2.cell(row=row, column=1, value="NOTES")
style_section(ws2, row, 6)

notes = [
    "Software platform cost = RM 0 (already built). RM 900/mo is pure margin.",
    "Infra: client pays ~RM 621/mo bundled, your actual cost ~RM 265/mo = ~RM 356/mo margin.",
    "Domain: client pays RM 250/yr lump, your cost ~RM 15/mo (RM 180/yr). RM 70/yr margin.",
    "WhatsApp OTP removed as separate line - absorbed into platform or handled separately.",
    "All 'Your Cost' cells are editable - adjust to match actual costs.",
]

for note in notes:
    row += 1
    ws2.merge_cells(f'A{row}:F{row}')
    ws2.cell(row=row, column=1, value=f"  \u2022 {note}")
    ws2.cell(row=row, column=1).font = Font(name='Calibri', size=10, color='808080')


# ═══════════════════════════════════════════════════════════
# SHEET 3: Option B - Self-Managed Infrastructure
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Option B - Self-Managed Infra")
ws3.sheet_properties.tabColor = "BF8F00"

ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 55
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 20

# Title
ws3.merge_cells('A1:D1')
ws3['A1'].value = "OPTION B: SELF-MANAGED INFRASTRUCTURE"
ws3['A1'].font = Font(name='Calibri', bold=True, size=14, color='BF8F00')
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 35

ws3.merge_cells('A2:D2')
ws3['A2'].value = "Client manages own cloud accounts & services | Lower monthly fee, higher operational effort"
ws3['A2'].font = Font(name='Calibri', italic=True, size=10, color='808080')
ws3['A2'].alignment = Alignment(horizontal='center')

# Headers
row = 4
ws3.cell(row=row, column=1, value="#")
ws3.cell(row=row, column=2, value="Description")
ws3.cell(row=row, column=3, value="Monthly (RM)")
ws3.cell(row=row, column=4, value="Annual (RM)")
style_header(ws3, row, 4)
ws3.row_dimensions[row].height = 25

# ─── A: Software License ───
row = 5
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1, value="A. SOFTWARE LICENSE (PAID TO YOU)")
style_section(ws3, row, 4)

lic_start = row + 1
license_items = [
    (1, "Admin Portal - Event management, agent/unit management, tier & rewards config, reports", 300),
    (2, "Agent Portal - Campaign browsing, link generation & sharing, partner management, rewards", 300),
    (3, "Public Pages - Registration, QR check-in display, OTP checkout, feedback", 300),
]

for i, (num, desc, price) in enumerate(license_items):
    row = lic_start + i
    ws3.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=2, value=desc)
    ws3.cell(row=row, column=3, value=price).number_format = currency_format
    ws3.cell(row=row, column=4).value = f"=C{row}*12"
    ws3.cell(row=row, column=4).number_format = currency_format
    style_row(ws3, row, 4)
    right_align(ws3, row, [3, 4])

row = lic_start + len(license_items)
ws3.cell(row=row, column=2, value="Subtotal - Software License")
ws3.cell(row=row, column=3).value = f"=SUM(C{lic_start}:C{row-1})"
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=4).value = f"=C{row}*12"
ws3.cell(row=row, column=4).number_format = currency_format
style_total(ws3, row, 4)
right_align(ws3, row, [3, 4])
sub3_a = row

# ─── B: Support & SLA ───
row += 1
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1, value="B. SUPPORT & SLA (PAID TO YOU)")
style_section(ws3, row, 4)

sup3_start = row + 1
support3_items = [
    (4, "SLA Support - 99.5% uptime guarantee (application layer only)", 250),
    (5, "Bug Fixes & Application Monitoring", 200),
    (6, "Minor Adjustments & Updates (up to 2 hours/month)", 200),
]

for i, (num, desc, price) in enumerate(support3_items):
    row = sup3_start + i
    ws3.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=2, value=desc)
    ws3.cell(row=row, column=3, value=price).number_format = currency_format
    ws3.cell(row=row, column=4).value = f"=C{row}*12"
    ws3.cell(row=row, column=4).number_format = currency_format
    style_row(ws3, row, 4)
    right_align(ws3, row, [3, 4])

row = sup3_start + len(support3_items)
ws3.cell(row=row, column=2, value="Subtotal - Support & SLA")
ws3.cell(row=row, column=3).value = f"=SUM(C{sup3_start}:C{row-1})"
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=4).value = f"=C{row}*12"
ws3.cell(row=row, column=4).number_format = currency_format
style_total(ws3, row, 4)
right_align(ws3, row, [3, 4])
sub3_b = row

# ─── C: Training ───
row += 1
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1, value="C. TRAINING & ENABLEMENT (PAID TO YOU)")
style_section(ws3, row, 4)

train3_start = row + 1
training3_items = [
    (7, "New Agent/Unit Onboarding Sessions (1 session/month, up to 20 pax)", 200),
]

for i, (num, desc, price) in enumerate(training3_items):
    row = train3_start + i
    ws3.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=2, value=desc)
    ws3.cell(row=row, column=3, value=price).number_format = currency_format
    ws3.cell(row=row, column=4).value = f"=C{row}*12"
    ws3.cell(row=row, column=4).number_format = currency_format
    style_row(ws3, row, 4)
    right_align(ws3, row, [3, 4])

row = train3_start + len(training3_items)
ws3.cell(row=row, column=2, value="Subtotal - Training & Enablement")
ws3.cell(row=row, column=3).value = f"=SUM(C{train3_start}:C{row-1})"
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=4).value = f"=C{row}*12"
ws3.cell(row=row, column=4).number_format = currency_format
style_total(ws3, row, 4)
right_align(ws3, row, [3, 4])
sub3_c = row

# ─── Subtotal Paid to You ───
row += 2
ws3.cell(row=row, column=2, value="SUBTOTAL - PAID TO YOU (Software + Support + Training)")
ws3.cell(row=row, column=3).value = f"=C{sub3_a}+C{sub3_b}+C{sub3_c}"
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=4).value = f"=C{row}*12"
ws3.cell(row=row, column=4).number_format = currency_format
style_total(ws3, row, 4, grand_total_fill)
right_align(ws3, row, [3, 4])
ws3.row_dimensions[row].height = 28
you_total_row = row

# ─── D: Client's Own Infra Costs ───
row += 2
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1, value="D. CLIENT-MANAGED INFRASTRUCTURE (PAID DIRECTLY BY CLIENT)")
style_section(ws3, row, 4)

client_infra_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

infra3_items = [
    (8, "Supabase Pro - Database, Auth, Edge Functions, Storage", 110, "supabase.com - Pro Plan USD 25/mo"),
    (9, "Render - Cloud Hosting (3 Static Sites, CI/CD)", 50, "render.com - Free tier or Starter USD ~12/mo"),
    (10, "Resend - Email Service (reminders, notifications)", 90, "resend.com - Pro Plan USD 20/mo"),
    (11, "Domain Registration & DNS (.com)", 15, "namecheap.com / cloudflare - ~USD 12/yr"),
]

infra3_start = row + 1
infra3_price_rows = []
for i, (num, desc, price, note) in enumerate(infra3_items):
    row = infra3_start + (i * 2)
    ws3.cell(row=row, column=1, value=num).alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=2, value=desc)
    ws3.cell(row=row, column=3, value=price).number_format = currency_format
    ws3.cell(row=row, column=4).value = f"=C{row}*12"
    ws3.cell(row=row, column=4).number_format = currency_format
    style_row(ws3, row, 4)
    right_align(ws3, row, [3, 4])
    for c in range(1, 5):
        ws3.cell(row=row, column=c).fill = client_infra_fill
    infra3_price_rows.append(row)
    # note sub-row
    row += 1
    ws3.cell(row=row, column=2, value=f"     Sign up: {note}")
    ws3.cell(row=row, column=2).font = Font(name='Calibri', size=9, italic=True, color='808080')
    style_row(ws3, row, 4)
    for c in range(1, 5):
        ws3.cell(row=row, column=c).fill = client_infra_fill

# Subtotal Client Infra
row += 1
ws3.cell(row=row, column=2, value="Subtotal - Client-Managed Infrastructure (est.)")
infra3_sum = "+".join([f"C{r}" for r in infra3_price_rows])
ws3.cell(row=row, column=3).value = f"={infra3_sum}"
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=4).value = f"=C{row}*12"
ws3.cell(row=row, column=4).number_format = currency_format
style_total(ws3, row, 4)
right_align(ws3, row, [3, 4])
for c in range(1, 5):
    ws3.cell(row=row, column=c).fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
    ws3.cell(row=row, column=c).font = Font(name='Calibri', bold=True, size=11, color='7F6000')
client_infra_total = row

# ─── TOTAL COST TO CLIENT ───
row += 2
ws3.cell(row=row, column=2, value="TOTAL MONTHLY COST TO CLIENT (You + Own Infra)")
ws3.cell(row=row, column=3).value = f"=C{you_total_row}+C{client_infra_total}"
ws3.cell(row=row, column=3).number_format = currency_format
ws3.cell(row=row, column=4).value = f"=C{row}*12"
ws3.cell(row=row, column=4).number_format = currency_format
style_total(ws3, row, 4, grand_total_fill)
right_align(ws3, row, [3, 4])
ws3.row_dimensions[row].height = 30
total_client_cost = row

# ─── E: Client Responsibilities ───
row += 2
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1, value="E. CLIENT RESPONSIBILITIES (SELF-MANAGED)")
style_section(ws3, row, 4)

warn_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
warn_font = Font(name='Calibri', size=10, color='C00000')

responsibilities = [
    "Sign up & maintain accounts: Supabase, Render, Resend",
    "Manage billing & payments for all cloud service subscriptions",
    "Monitor database storage and email quota usage",
    "Handle Supabase database backups & disaster recovery",
    "Manage domain DNS settings, SSL certificate renewals",
    "Coordinate with cloud providers for any service outages",
    "Upgrade service plans as usage grows",
    "Provide service credentials/API keys to developer for setup & deployments",
    "Infrastructure uptime NOT covered under SLA (application-level SLA only)",
]

for resp in responsibilities:
    row += 1
    ws3.cell(row=row, column=1, value="  !").font = Font(name='Calibri', bold=True, color='C00000')
    ws3.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws3.merge_cells(f'B{row}:D{row}')
    ws3.cell(row=row, column=2, value=resp).font = warn_font
    style_row(ws3, row, 4)
    for c in range(1, 5):
        ws3.cell(row=row, column=c).fill = warn_fill

# ─── Comparison ───
row += 2
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1, value="OPTION A vs OPTION B COMPARISON")
style_section(ws3, row, 4)

row += 1
ws3.cell(row=row, column=2, value="")
ws3.cell(row=row, column=3, value="Option A\n(Managed SaaS)")
ws3.cell(row=row, column=4, value="Option B\n(Self-Managed)")
style_header(ws3, row, 4)
ws3.row_dimensions[row].height = 35

compare_items = [
    ("Monthly Fee (to you)", "RM 2,600", f"=C{you_total_row}", False),
    ("Client's Own Infra Cost", "RM 0", f"=C{client_infra_total}", False),
    ("Total Monthly Cost to Client", "RM 2,600", f"=C{total_client_cost}", True),
    ("Annual Cost to Client", "RM 31,200", f"=C{total_client_cost}*12", True),
    ("", "", "", False),
    ("Infrastructure Management", "Fully managed by you", "Client manages own", False),
    ("Cloud Account Setup", "Not required", "Client must sign up", False),
    ("Service Billing", "Single invoice from you", "Multiple vendor bills", False),
    ("Database Backups", "Included & automated", "Client's responsibility", False),
    ("Scaling & Upgrades", "You handle seamlessly", "Client coordinates", False),
    ("SLA Coverage", "Full stack (app + infra)", "App layer only", False),
    ("Troubleshooting", "Single point of contact", "You + cloud vendors", False),
    ("Operational Effort", "Zero for client", "Moderate (est. 2-4 hrs/mo)", False),
]

for label, opt_a, opt_b, is_highlight in compare_items:
    row += 1
    ws3.cell(row=row, column=2, value=label).font = bold_font if is_highlight else normal_font
    ws3.cell(row=row, column=3, value=opt_a)
    if isinstance(opt_b, str) and not opt_b.startswith("="):
        ws3.cell(row=row, column=4, value=opt_b)
    else:
        ws3.cell(row=row, column=4).value = opt_b
        ws3.cell(row=row, column=4).number_format = currency_format
    style_row(ws3, row, 4)
    ws3.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    ws3.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    if is_highlight:
        for c in range(1, 5):
            ws3.cell(row=row, column=c).font = Font(name='Calibri', bold=True, size=11)
        ws3.cell(row=row, column=3).fill = margin_good_fill
        ws3.cell(row=row, column=3).font = Font(name='Calibri', bold=True, size=11, color='375623')

# ─── Recommendation ───
row += 2
ws3.merge_cells(f'A{row}:D{row}')
ws3.cell(row=row, column=1, value="RECOMMENDATION")
style_section(ws3, row, 4)

rec_notes = [
    "Option A (Managed SaaS) is RECOMMENDED:",
    "",
    "1. Single monthly invoice - no managing multiple cloud vendor bills",
    "2. Full-stack SLA - infrastructure issues covered, faster resolution",
    "3. Zero operational overhead - no cloud admin knowledge required",
    "4. Seamless scaling - upgrades handled transparently as usage grows",
    "5. Lower total effort - client saves estimated 2-4 hours/month on infra management",
    f"6. Price difference is minimal for significantly less hassle",
    "",
    "Option B is suitable if client has in-house IT team and prefers direct cloud control.",
]

for note in rec_notes:
    row += 1
    ws3.merge_cells(f'A{row}:D{row}')
    ws3.cell(row=row, column=1, value=f"  {note}")
    ws3.cell(row=row, column=1).font = Font(name='Calibri', size=10, color='1F4E79')


# ─── Print settings ───
for ws in [ws1, ws2, ws3]:
    ws.print_area = ws.dimensions
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

# Save
output_path = "/Users/paullee/Documents/project/martin/DATA/docs/Agent-Onboarding-System-Costing.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
