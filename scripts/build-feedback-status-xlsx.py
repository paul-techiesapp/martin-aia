#!/usr/bin/env python3
"""Generate the feedback-changes status workbook (Feedback Changes.pdf round 1).

Usage: python3 scripts/build-feedback-status-xlsx.py
Output: docs/feedback/2026-06-30-feedback-changes-status.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "docs/feedback/2026-06-30-feedback-changes-status.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1A1942")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1A1942")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
DONE_FILL = PatternFill("solid", fgColor="E6F4EA")

# (#, Page, Feedback item, Area, What changed, Files / Migrations, Status)
FEEDBACK = [
    ("1", "1", "Partnerships: remove the Pool (RM) limit — “no need to set limit here”",
     "Admin · Partnerships",
     "Removed the Pool (RM) column and the Merchant/Customer split column from the list, and the Gift Pool + Merchant Share fields from the create/edit form. A partnership is now just Name + Logo + Status.",
     "apps/admin-portal/src/pages/merchants/MerchantList.tsx, MerchantDetail.tsx; useMerchants.ts",
     "Done (staging)"),
    ("2", "1", "Customer payout is standard — 10% of total Car Insurance Renewal",
     "Admin · Partnerships / Renewal",
     "Customer gift = editable rate (default 10%) × the renewal premium, captured when admin confirms a renewal. Merchant settlement (payable) equals that gift. Rate set in Settings.",
     "migration 20260630000001 (confirm_vehicle_renewal rewrite); Settings.tsx (gift rate); EnquiryDetail.tsx (renewal dialog)",
     "Done (staging)"),
    ("3", "2 & 4", "Enquiries view default sort (filterable): Units > Agents > Partners > Status > Expiration > Received",
     "Admin + Agent · Enquiries",
     "Default multi-key sort in that exact order (nulls last) plus Unit / Agent / Partner / Status filter dropdowns. Applied to the admin Inbox and the agent My Enquiries list.",
     "admin: EnquiryList.tsx, enquirySort.ts, useEnquiries.ts; agent: MyEnquiries.tsx, myEnquiriesSort.ts",
     "Done"),
    ("4", "2 & 4", "Need to be able to download the enquiries report",
     "Admin + Agent · Enquiries",
     "“Download report” button exports an .xlsx (one row per car: Unit, Agent, Partner, Customer, Phone, Email, Plate, Expiry, Road Tax, statuses, Received), honoring the active filters + sort.",
     "shared-ui buildEnquiriesWorkbook; EnquiryList.tsx; MyEnquiries.tsx",
     "Done"),
    ("5", "3", "Assign to partner — confirmed by each specific quotation upon successful renewal",
     "Admin + Agent · Enquiries",
     "Partner now binds per car at renewal: admin picks/confirms the partner in the renewal-confirm dialog (pre-filled from the agent’s enquiry-level suggestion). Per-car merchant stored on the vehicle.",
     "migration 20260630000001 (enquiry_vehicles.merchant_id); EnquiryDetail.tsx",
     "Done (staging)"),
    ("6", "3", "Add “GET QUOTE” button → auto email to Admin (agent & unit info + customer & vehicle details)",
     "Agent · Enquiries + Edge function",
     "Per-car Get Quote button emails the admin with agent+unit (incl. unit admin), customer and vehicle details. Ownership-guarded, idempotent (no duplicate emails), stamps quote_requested_at.",
     "agent MyEnquiries.tsx, useRequestQuote.ts; supabase/functions/send-quote-request",
     "Done (fn deployed to staging)"),
    ("7", "5", "Master Admin: download report of successful renewal cases — sort/filter by Partnership, Units, Agent, Timeline, Value",
     "Admin · Reports",
     "New “Renewals” tab: filter by Partner / Unit / Agent / date range, sort by value, .xlsx download. Gift/settlement figures read the actual minted ledger amounts (not re-derived from the live rate).",
     "Reports.tsx (Renewals tab); useRenewalReport.ts; shared-ui buildRenewalsWorkbook",
     "Done"),
    ("8", "5 & 6", "Inquiry form: add T&C agreement + accept T&C (PDPA text)",
     "Public form + Admin Settings",
     "Scrollable PDPA Terms block + required “I accept” checkbox; form blocks until accepted. T&C body is admin-editable in Settings and seeded with the page-6 PDPA clause.",
     "public Enquiry.tsx, useEnquiryFormSettings.ts; admin Settings.tsx; migration 20260630000002",
     "Done (staging)"),
    ("9", "5", "Inquiry form: add Header & Footer design we can customize",
     "Public form + Admin Settings",
     "Header (logo, title, subtitle) and footer text are admin-editable via Settings and stored in system_settings.enquiry_form; the public form reads them live (merchant branding still overlays branch links).",
     "admin Settings.tsx, useSystemSettings.ts; public Enquiry.tsx; migration 20260630000002",
     "Done (staging)"),
    ("10", "5", "Inquiry form: all fields mandatory including Covernote/Geran upload",
     "Public form",
     "Email is now required + validated; each car must attach ≥1 Covernote/Geran document (submit is blocked otherwise); uploader relabeled “Covernote / Geran (required)”.",
     "public Enquiry.tsx",
     "Done"),
    ("11", "5", "Inquiry form: add option for Road Tax renewal",
     "Public form + DB",
     "Per-car required “Road Tax Renewal? Yes / No” choice, persisted with the enquiry and shown in exports, the Get Quote email, and admin views.",
     "public Enquiry.tsx; migrations 20260630000001 (column) + 20260630000003 (submit_enquiry)",
     "Done (staging)"),
    ("12", "Sidebar (verbal)", "Segregate the sidebar into 2 parts: Events and Partnership",
     "Admin + Agent · Navigation",
     "Both sidebars are grouped under EVENTS and PARTNERSHIP section headers (Dashboard/Settings/Account stay ungrouped). Applied to every agent role variant.",
     "admin Layout.tsx; agent Layout.tsx",
     "Done"),
]

# Verified review findings (all CONFIRMED by adversarial verification, all fixed)
REVIEW = [
    ("R1", "Medium", "Renewals report recomputed gift from the live rate instead of the minted ledger → wrong figures if the rate changes.",
     "Report now reads gifts.value_amount / merchant_settlements.amount (source of truth), falling back to recompute only when no ledger row exists.",
     "useRenewalReport.ts, Reports.tsx", "Fixed"),
    ("R2", "Medium", "send-quote-request authorized any logged-in user — IDOR: could email/stamp enquiries they don’t own.",
     "Caller must own the enquiry (or be admin/service-role); vehicle must belong to the enquiry.",
     "supabase/functions/send-quote-request/index.ts", "Fixed"),
    ("R3", "Low", "confirm_vehicle_renewal lacked a submitted|quoted guard — a direct re-confirm could desync ledger vs vehicle.",
     "Added status guard rejecting already-renewed/lost vehicles.",
     "migration 20260630000001", "Fixed"),
    ("R4", "Low", "Renewals date filter compared a date string to UTC timestamps — off by the +8h SGT offset.",
     "Bounded the range in Singapore time (+08:00).",
     "useRenewalReport.ts", "Fixed"),
    ("R5", "Low", "Get Quote showed a success toast even when the function gracefully skipped (no admin email / no Resend key).",
     "UI now branches on skipped / alreadyRequested and shows the right message.",
     "agent MyEnquiries.tsx", "Fixed"),
    ("R6", "Low", "Vehicle was stamped after the email and a stamp failure was swallowed → possible duplicate admin emails.",
     "Atomic claim before send + rollback (and 502) on send failure; idempotent on repeat.",
     "supabase/functions/send-quote-request/index.ts", "Fixed"),
    ("R7", "Low", "Per-car partner wasn’t reflected in the enquiry exports (used the enquiry-level partner for every car).",
     "Exports now use the per-vehicle partner, falling back to the enquiry-level suggestion.",
     "admin useEnquiries.ts + enquirySort.ts; agent useMyEnquiries.ts + MyEnquiries.tsx", "Fixed"),
]

NOTES = [
    "Deploy target: STAGING Supabase only (BOP Website Staging, lyjdlietzmmejrxjvwgp). Production untouched.",
    "Migrations applied to staging: 20260630000001 (gift rate + per-car renewal cols + confirm_vehicle_renewal), 20260630000002 (enquiry-form settings + admin email), 20260630000003 (submit_enquiry road tax).",
    "Edge function send-quote-request deployed to staging (verify_jwt on).",
    "ACTION NEEDED: set the Admin Notification Email in admin Settings so Get Quote emails actually send (currently blank → the flow gracefully skips).",
    "ACTION NEEDED: ensure RESEND_API_KEY is set as a staging edge-function secret for Get Quote emails to send.",
    "Frontend staging sites (racc-*-staging on Render) deploy automatically when the feat/merchant-partnership branch is pushed.",
    "All work is on branch feat/merchant-partnership. All three apps build clean; packages typecheck clean. (Repo has no unit-test runner.)",
    "A 6-dimension adversarial code review (13 agents) ran over the diff; all 7 confirmed findings above were fixed and re-built.",
]


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_TOP
        cell.border = BORDER


def main():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()

    # Sheet 1 — Feedback status
    ws = wb.active
    ws.title = "Feedback Status"
    headers = ["#", "Page", "Feedback Item", "Area", "What Changed", "Files / Migrations", "Status"]
    widths = [5, 12, 42, 24, 60, 50, 20]
    ws.append(headers)
    style_header(ws, len(headers))
    for row in FEEDBACK:
        ws.append(list(row))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        status_cell = ws.cell(row=r, column=len(headers))
        if str(status_cell.value).startswith("Done"):
            status_cell.fill = DONE_FILL
    ws.freeze_panes = "A2"

    # Sheet 2 — QA review fixes
    ws2 = wb.create_sheet("QA Review Fixes")
    h2 = ["ID", "Severity", "Finding", "Fix Applied", "Files", "Status"]
    w2 = [6, 10, 60, 55, 45, 12]
    ws2.append(h2)
    style_header(ws2, len(h2))
    for row in REVIEW:
        ws2.append(list(row))
    for i, w in enumerate(w2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, ws2.max_row + 1):
        for c in range(1, len(h2) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.alignment = WRAP_TOP
            cell.border = BORDER
        ws2.cell(row=r, column=len(h2)).fill = DONE_FILL
    ws2.freeze_panes = "A2"

    # Sheet 3 — Notes / deployment
    ws3 = wb.create_sheet("Notes & Deployment")
    ws3["A1"] = "Notes & Deployment"
    ws3["A1"].font = TITLE_FONT
    ws3.column_dimensions["A"].width = 120
    for i, note in enumerate(NOTES, start=3):
        c = ws3.cell(row=i, column=1, value="•  " + note)
        c.alignment = WRAP_TOP

    wb.save(OUT)
    print(f"Wrote {OUT} ({os.path.getsize(OUT)} bytes)")


if __name__ == "__main__":
    main()
